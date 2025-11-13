"""
ILAS Stable Admin Integration (Non-Recursive, Audit-Safe)
--------------------------------------------------------
Implements:
• Full Book management (CRUD + Bulk Upload)
• Book Transactions (Issue / Return)
• AuditLog view (read-only)
• Signal-safe admin operations
"""

import os
import zipfile
import openpyxl
from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django.core.files.base import ContentFile
from django.db import transaction

from .models import Book, BookTransaction, AuditLog
from .serializers import BookTransactionSerializer, BulkBookImportSerializer
from .models import create_audit


# ----------------------------------------------------------------------
# INLINE CONFIG
# ----------------------------------------------------------------------
class BookTransactionInline(admin.TabularInline):
    model = BookTransaction
    extra = 0
    readonly_fields = (
        "txn_type", "member", "actor",
        "issue_date", "due_date", "return_date", "fine_amount",
    )
    can_delete = False
    ordering = ("-issue_date",)


# ----------------------------------------------------------------------
# BOOK ADMIN
# ----------------------------------------------------------------------
@admin.register(Book)
class BookAdmin(admin.ModelAdmin):
    list_display = ("book_code", "title", "author", "status", "issued_to", "shelf_location", "created_at")
    list_filter = ("status", "category", "language")
    search_fields = ("book_code", "title", "author", "isbn")
    readonly_fields = ("book_code", "uid", "created_at", "updated_at", "last_modified_by")
    ordering = ("-created_at",)
    inlines = [BookTransactionInline]
    change_list_template = "admin/library/book_change_list.html"

    fieldsets = (
        ("Book Metadata", {
            "fields": (
                "book_code", "title", "subtitle", "author", "publisher", "edition",
                "publication_year", "isbn", "language", "category",
                "keywords", "description", "cover_image",
            )
        }),
        ("Physical Info", {
            "fields": (
                "accession_no", "shelf_location", "condition", "book_cost",
                "vendor_name", "source", "library_section", "dewey_decimal",
                "cataloger", "remarks",
            )
        }),
        ("System Fields", {
            "fields": ("status", "issued_to", "last_modified_by", "created_at", "updated_at")
        }),
    )

    # ------------------------------------------------------------------
    # CUSTOM ADMIN URLS
    # ------------------------------------------------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("bulk-upload/", self.admin_site.admin_view(self.bulk_upload_view), name="library-book-bulk-upload"),
            path("issue-book/", self.admin_site.admin_view(self.issue_book_view), name="library-book-issue-book"),
            path("return-book/", self.admin_site.admin_view(self.return_book_view), name="library-book-return-book"),
        ]
        return custom_urls + urls

        # ------------------------------------------------------------------
    # BULK UPLOAD (Excel + ZIP)
    # ------------------------------------------------------------------
    def bulk_upload_view(self, request):
        """Handle Excel + ZIP bulk import with audit suppression."""
        if request.method != "POST":
            return render(request, "admin/library/book_bulk_upload.html")

        excel_file = request.FILES.get("file")
        images_zip = request.FILES.get("images")

        if not excel_file:
            messages.error(request, "Excel file is required.")
            return redirect("..")

        # ----------------------------------------------------------
        # Step 1: Load images from ZIP (if provided)
        # ----------------------------------------------------------
        images_map = {}
        if images_zip:
            try:
                with zipfile.ZipFile(images_zip) as z:
                    for name in z.namelist():
                        base = name.split("/")[-1].lower()
                        if base.endswith((".jpg", ".jpeg", ".png")):
                            images_map[base] = z.read(name)
            except zipfile.BadZipFile:
                messages.warning(request, "⚠️ Invalid ZIP file uploaded. Skipping images.")

        # ----------------------------------------------------------
        # Step 2: Parse Excel
        # ----------------------------------------------------------
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

            # Deduplicate header columns (avoid collisions)
            seen = {}
            for i, h in enumerate(header):
                if h in seen:
                    seen[h] += 1
                    header[i] = f"{h}_{seen[h]}"
                else:
                    seen[h] = 1

        except Exception as e:
            messages.error(request, f"Excel error: {e}")
            return redirect("..")

        # ----------------------------------------------------------
        # Step 3: Process rows
        # ----------------------------------------------------------
        created, errors = 0, []
        try:
            with transaction.atomic():
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(v in (None, "") for v in row):
                        continue

                    data = dict(zip(header, row))
                    serializer = BulkBookImportSerializer(data=data)
                    try:
                        serializer.is_valid(raise_exception=True)
                        book = Book(**serializer.validated_data, last_modified_by=request.user)
                        book._suppress_audit = True
                        book.save()

                        # Optional cover assignment
                        for key in [
                            f"{(book.isbn or '').strip().lower()}.jpg",
                            f"{(book.title or '').strip().lower()}.jpg",
                        ]:
                            if key in images_map:
                                book.cover_image.save(f"{book.book_code}.jpg", ContentFile(images_map[key]), save=True)
                                break

                        created += 1
                    except Exception as e:
                        errors.append(f"Row {i}: {e}")
                        continue

            # ------------------------------------------------------
            # Step 4: Single bulk audit entry
            # ------------------------------------------------------
            create_audit(
                actor=request.user,
                action=AuditLog.ACTION_BULK_UPLOAD,
                target_type="Book",
                target_id="BulkImport",
                new_values={"books_created": created},
                remarks=f"Bulk uploaded {created} books via admin interface",
                source="admin-ui",
            )

        except Exception as e:
            messages.error(request, f"Bulk upload failed: {e}")
            return redirect("..")

        # ----------------------------------------------------------
        # Step 5: Feedback
        # ----------------------------------------------------------
        if created:
            messages.success(request, f"✅ {created} books uploaded successfully.")
        if errors:
            sample = errors[:3]
            messages.warning(request, f"⚠️ {len(errors)} rows failed. Example: {sample}")
        return redirect("..")


    # ------------------------------------------------------------------
    # ISSUE / RETURN (Validated, Audit-Safe)
    # ------------------------------------------------------------------
    def issue_book_view(self, request):
        """Admin interface for issuing a book (with validation)."""
        from django.contrib.auth import get_user_model
        User = get_user_model()

        if request.method == "POST":
            book_id = request.POST.get("book_id")
            member_id = request.POST.get("member_id")
            remarks = request.POST.get("remarks", "").strip()
            book = Book.objects.filter(id=book_id).first()
            member = User.objects.filter(id=member_id).first()

            if not book or not member:
                messages.error(request, "Invalid book or member.")
                return redirect("..")

            # inside admin view function before calling book.mark_issued
            if not book.can_be_issued():
                messages.error(request, f"Book {book.book_code} is not available for issue.")
                return redirect(request.path)  # or appropriate view
            # Prefer serializer:
            serializer = BookTransactionSerializer(data={
                "book": book.id, "member": member.id, "txn_type": BookTransaction.TYPE_ISSUE
            }, context={"request": request})
            if not serializer.is_valid():
                messages.error(request, f"Cannot issue: {serializer.errors}")
                return redirect(request.path)
            # txn = serializer.save()
            # messages.success(request, f"Issued {book.book_code} to {member.username}")

            # active_txn = BookTransaction.objects.filter(
            #     book=book, txn_type=BookTransaction.TYPE_ISSUE, is_active=True
            # ).first()
            # if active_txn:
            #     messages.error(request, f"Book {book.book_code} is already issued to {active_txn.member}.")
            #     return redirect("..")
            try:
                with transaction.atomic():
                    serializer.is_valid(raise_exception=True)
                    txn = serializer.save()
                    messages.success(request, f"✅ {book.book_code} issued to {member.username}.")
            except Exception as e:
                messages.error(request, f"Issue failed: {e}")
            return redirect("..")
        
        # --- GET: show available books ---
        books = Book.objects.filter(status=Book.STATUS_AVAILABLE)
        context = {"books": books[:200]}  # increased limit
        return render(request, "admin/library/issue_book.html", context)


    def return_book_view(self, request):
        """Admin interface for returning a book (same-member validation)."""
        from django.contrib.auth import get_user_model
        User = get_user_model()

        if request.method == "POST":
            book_id = request.POST.get("book_id")
            member_id = request.POST.get("member_id")
            remarks = request.POST.get("remarks", "").strip()
            book = Book.objects.filter(id=book_id).first()
            member = User.objects.filter(id=member_id).first()

            if not book or not member:
                messages.error(request, "Invalid book or member.")
                return redirect("..")

            active_txn = BookTransaction.objects.filter(
                book=book, txn_type=BookTransaction.TYPE_ISSUE, is_active=True
            ).first()
            if not active_txn:
                messages.error(request, f"No active issue found for {book.book_code}.")
                return redirect("..")

            # --- Enforce same-member return ---
            if str(active_txn.member.id) != str(member.id):
                messages.error(request, f"Book {book.book_code} was issued to {active_txn.member.username}.")
                return redirect("..")

            try:
                with transaction.atomic():
                    book.mark_returned(actor=request.user, remarks=remarks)
                    messages.success(request, f"✅ {book.book_code} returned successfully.")
            except ValueError as e:
                messages.error(request, f"Return failed: {e}")
            except Exception as e:
                messages.error(request, f"Unexpected error: {e}")
            return redirect("..")

        # --- GET: show issued books ---
        issued_books = Book.objects.filter(status=Book.STATUS_ISSUED)
        context = {"books": issued_books[:200]}
        return render(request, "admin/library/return_book.html", context)


    # ------------------------------------------------------------------
    # SAVE + DELETE (AUDIT-SAFE)
    # ------------------------------------------------------------------
    def save_model(self, request, obj, form, change):
        """Save book safely without duplicate audits."""
        obj.last_modified_by = request.user
        obj._suppress_audit = False  # allow signal audit (handled once)
        super().save_model(request, obj, form, change)

    def delete_model(self, request, obj):
        """Delete book safely — signal handles audit logging."""
        obj._suppress_audit = False
        super().delete_model(request, obj)


# ----------------------------------------------------------------------
# BOOK TRANSACTION ADMIN
# ----------------------------------------------------------------------
@admin.register(BookTransaction)
class BookTransactionAdmin(admin.ModelAdmin):
    list_display = ("id", "book", "member", "txn_type", "issue_date", "due_date", "return_date", "fine_amount", "is_active")
    list_filter = ("txn_type", "is_active")
    search_fields = ("book__book_code", "member__username")
    readonly_fields = ("actor", "issue_date", "due_date", "return_date", "fine_amount")
    ordering = ("-created_at",)

    def save_model(self, request, obj, form, change):
        """Audit-safe transaction save."""
        obj.actor = request.user
        obj._suppress_audit = False
        super().save_model(request, obj, form, change)
        # No manual create_audit here — handled by signal


# ----------------------------------------------------------------------
# AUDIT LOG ADMIN (Read-Only)
# ----------------------------------------------------------------------
@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display = ("actor", "action", "target_type", "target_id", "timestamp", "source")
    list_filter = ("action", "source", "target_type")
    search_fields = ("actor__username", "target_id", "action")
    ordering = ("-timestamp",)
    readonly_fields = ("actor", "action", "target_type", "target_id", "old_values", "new_values", "remarks", "timestamp")

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False
